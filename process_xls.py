import openpyxl
with open("./hnk_2015_new.xlsx", "rb") as excel:
    wb=openpyxl.load_workbook(excel)
    sheet=wb["Locality_data"]

    counties=[]
    for cell in sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        county=cell[0]
        if(county!=None and county not in counties):
            counties.append(county)

    nr=0
    county_map={}
    for county in counties:
        county_id=f"c_{nr}"
        county_map[county]=county_id
        print(f"create ({county_id}:County {{name: '{county}'}})")
        nr+=1



    l_nr=0
    for cell in sheet.iter_rows(min_row=2, max_row=3179, values_only=True):
        locality_id=f"l_{l_nr}"
        l_nr+=1
        locality_name=cell[0]
        locality_type=cell[2]
        locality_county=cell[3]
        locality_size=cell[9]
        locality_population=cell[10]
        locality_apartment_count=cell[11]

        county_id=county_map[locality_county]

        print(f"create({locality_id}: Locality {{name: '{locality_name}', type: '{locality_type}', size: {locality_size}, population: {locality_population}, apartments: {locality_apartment_count}}})")
        print(f",({locality_id})-[:LOCATED_IN]->({county_id})")


# import pandas as ps
# import json

# data=ps.read_excel("./hnk_2015_new.xlsx", index_col=None, sheet_name="Locality_data", engine='openpyxl')


